import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import openpyxl
import os
import shutil
from database.conexion import db
from database.queries import Queries

class VentanaExportarAsistencia:
    """Ventana para exportar datos de asistencia/nómina usando plantilla existente"""
    
    def __init__(self, parent):
        self.parent = parent
        
        self.ventana = tk.Toplevel(parent)
        self.ventana.title("Exportar Datos de Asistencia / Nómina")
        self.ventana.geometry("650x650")
        self.ventana.transient(parent)
        self.ventana.grab_set()
        
        self.centrar_ventana()
        self.crear_interfaz()
    
    def centrar_ventana(self):
        self.ventana.update_idletasks()
        ancho = 650
        alto = 650
        x = (self.ventana.winfo_screenwidth() // 2) - (ancho // 2)
        y = (self.ventana.winfo_screenheight() // 2) - (alto // 2)
        self.ventana.geometry(f'{ancho}x{alto}+{x}+{y}')
    
    def crear_interfaz(self):
        # Frame principal
        main_frame = ttk.Frame(self.ventana, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        titulo = ttk.Label(main_frame, text="Exportar Datos de Asistencia / Nómina",
                          font=('Helvetica', 14, 'bold'))
        titulo.pack(pady=(0, 10))
        
        subtitulo = ttk.Label(main_frame, text="Complete los parámetros y seleccione su plantilla",
                             font=('Helvetica', 10), foreground='gray')
        subtitulo.pack(pady=(0, 20))
        
        # Frame para parámetros
        params_frame = ttk.LabelFrame(main_frame, text="Parámetros del Reporte", padding="15")
        params_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Fecha de inicio de periodo
        ttk.Label(params_frame, text="Fecha de Inicio de Periodo:", font=('Helvetica', 10, 'bold')).grid(row=0, column=0, sticky='w', pady=8)
        self.fecha_inicio = tk.StringVar()
        entry_fecha = ttk.Entry(params_frame, textvariable=self.fecha_inicio, width=20)
        entry_fecha.grid(row=0, column=1, sticky='w', pady=8, padx=(10, 0))
        ttk.Label(params_frame, text="(Formato: YYYY-MM-DD)", font=('Helvetica', 8), foreground='gray').grid(row=0, column=2, sticky='w', pady=8, padx=(10, 0))
        
        # Tasa de cambio del mes
        ttk.Label(params_frame, text="Tasa de Cambio del Mes (Bs/USD):", font=('Helvetica', 10, 'bold')).grid(row=1, column=0, sticky='w', pady=8)
        self.tasa_cambio = tk.StringVar()
        entry_tasa = ttk.Entry(params_frame, textvariable=self.tasa_cambio, width=20)
        entry_tasa.grid(row=1, column=1, sticky='w', pady=8, padx=(10, 0))
        ttk.Label(params_frame, text="(Ejemplo: 40.50)", font=('Helvetica', 8), foreground='gray').grid(row=1, column=2, sticky='w', pady=8, padx=(10, 0))
        
        # Frame para plantilla
        plantilla_frame = ttk.LabelFrame(main_frame, text="Plantilla de Excel", padding="10")
        plantilla_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.ruta_plantilla = tk.StringVar()
        ttk.Label(plantilla_frame, text="Seleccione su plantilla:").grid(row=0, column=0, sticky='w', padx=5)
        entry_plantilla = ttk.Entry(plantilla_frame, textvariable=self.ruta_plantilla, width=45)
        entry_plantilla.grid(row=0, column=1, sticky='w', padx=5)
        
        btn_buscar = tk.Button(plantilla_frame, text="📂 Buscar", command=self.buscar_plantilla,
                              bg="#3498db", fg="white", cursor='hand2')
        btn_buscar.grid(row=0, column=2, padx=5)
        
        # Información de empleados
        info_frame = ttk.Frame(plantilla_frame)
        info_frame.grid(row=1, column=0, columnspan=3, sticky='ew', pady=10)
        
        total_empleados = Queries.contar_empleados()
        ttk.Label(info_frame, text=f"✅ Empleados activos disponibles: {total_empleados}", 
                 font=('Helvetica', 9), foreground='green').pack()
        
        # Botones de acción
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=10)
        
        btn_generar = tk.Button(action_frame, text="📊 Generar Reporte", command=self.generar_reporte,
                               bg="#2ecc71", fg="white", font=('Helvetica', 11),
                               padx=20, pady=8, cursor='hand2')
        btn_generar.pack(side=tk.LEFT, padx=5)
        
        btn_cancelar = tk.Button(action_frame, text="❌ Cancelar", command=self.ventana.destroy,
                                bg="#e74c3c", fg="white", cursor='hand2')
        btn_cancelar.pack(side=tk.RIGHT, padx=5)
    
    def buscar_plantilla(self):
        """Busca el archivo de plantilla de Excel"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar plantilla de Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if archivo:
            self.ruta_plantilla.set(archivo)
    
    def validar_datos(self):
        """Valida los datos ingresados"""
        # Validar fecha de inicio
        fecha = self.fecha_inicio.get().strip()
        if not fecha:
            messagebox.showerror("Error", "La fecha de inicio de periodo es obligatoria")
            return False
        
        try:
            datetime.strptime(fecha, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Error", "Formato de fecha inválido. Use YYYY-MM-DD")
            return False
        
        # Validar tasa de cambio
        tasa = self.tasa_cambio.get().strip()
        if not tasa:
            messagebox.showerror("Error", "La tasa de cambio es obligatoria")
            return False
        
        try:
            tasa_float = float(tasa)
            if tasa_float <= 0:
                messagebox.showerror("Error", "La tasa de cambio debe ser mayor a cero")
                return False
        except ValueError:
            messagebox.showerror("Error", "La tasa de cambio debe ser un número válido")
            return False
        
        # Validar plantilla
        plantilla = self.ruta_plantilla.get().strip()
        if not plantilla:
            messagebox.showerror("Error", "Debe seleccionar una plantilla de Excel")
            return False
        
        if not os.path.exists(plantilla):
            messagebox.showerror("Error", "El archivo de plantilla no existe")
            return False
        
        return True
    
    def obtener_datos_empleados(self):
        """Obtiene los datos de los empleados ACTIVOS para el reporte"""
        query = """
            SELECT 
                ROW_NUMBER() OVER (ORDER BY e.nombres_apellidos) as numero,
                e.numero_cuenta as cuenta,
                b.nombre_banco as tipo_banco,
                e.nombres_apellidos as nombre_apellido,
                e.cedula,
                tn.nombre_nomina as nomina,
                tp.nombre_tipo as tipo_personal,
                c.nombre_cargo as cargo
            FROM empleados e
            LEFT JOIN tipos_nomina tn ON e.id_tipo_nomina = tn.id
            LEFT JOIN tipos_personal tp ON e.id_tipo_personal = tp.id
            LEFT JOIN cargos c ON e.id_cargo = c.id
            LEFT JOIN bancos b ON e.id_banco = b.id
            WHERE e.estatus = 'ACTIVO'
            ORDER BY e.nombres_apellidos
        """
        return db.execute_query(query)
    
    def generar_reporte(self):
        """Genera el reporte usando la plantilla existente"""
        if not self.validar_datos():
            return
        
        # Obtener datos de empleados
        datos = self.obtener_datos_empleados()
        
        if not datos:
            messagebox.showwarning("Advertencia", "No hay empleados activos para exportar")
            return
        
        # Seleccionar archivo de destino
        fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_destino = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"reporte_asistencia_{fecha_archivo}.xlsx"
        )
        
        if not archivo_destino:
            return
        
        try:
            # Copiar la plantilla al destino
            plantilla = self.ruta_plantilla.get().strip()
            shutil.copy2(plantilla, archivo_destino)
            
            # Abrir el libro copiado
            wb = openpyxl.load_workbook(archivo_destino)
            
            # Seleccionar la hoja activa (o la primera hoja)
            ws = wb.active
            
            # === ESCRIBIR PARÁMETROS EN LAS CELDAS ESPECÍFICAS ===
            # Fecha en celda O1
            ws['O1'] = self.fecha_inicio.get().strip()
            
            # Tasa de cambio en celda Q1
            ws['Q1'] = float(self.tasa_cambio.get().strip())
            
            # === ESCRIBIR DATOS DE EMPLEADOS DESDE A5 HASTA H5 ===
            # Las columnas son: A=N°, B=Cuenta, C=Tipo Banco, D=Nombre, E=Cédula, F=Nómina, G=Tipo Personal, H=Cargo
            for idx, row in enumerate(datos, start=5):
                ws.cell(row=idx, column=1, value=row[0])   # N° (columna A)
                ws.cell(row=idx, column=2, value=row[1])   # Cuenta (columna B)
                ws.cell(row=idx, column=3, value=row[2])   # Tipo de banco (columna C)
                ws.cell(row=idx, column=4, value=row[3])   # Nombre y apellido (columna D)
                ws.cell(row=idx, column=5, value=row[4])   # Cédula (columna E)
                ws.cell(row=idx, column=6, value=row[5])   # Nómina (columna F)
                ws.cell(row=idx, column=7, value=row[6])   # Tipo de personal (columna G)
                ws.cell(row=idx, column=8, value=row[7])   # Cargo (columna H)
            
            # Guardar el archivo
            wb.save(archivo_destino)
            
            messagebox.showinfo("Éxito", 
                f"✅ Reporte generado correctamente!\n\n"
                f"📁 Archivo: {archivo_destino}\n"
                f"📊 Empleados procesados: {len(datos)}\n"
                f"📅 Fecha período: {self.fecha_inicio.get()}\n"
                f"💵 Tasa de cambio: {self.tasa_cambio.get()} Bs/USD")
            
            self.ventana.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar el reporte:\n{str(e)}")